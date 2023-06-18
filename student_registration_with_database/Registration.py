from tkinter import *
from tkinter import filedialog, messagebox
from tkinter.ttk import Combobox
from datetime import date
from PIL import ImageTk, Image
import openpyxl, xlrd
from openpyxl import workbook
#import xlrd
import pathlib
import os

background = '#06283D'
framebg = '#EDEDED'
framefg = '#06283D'

master = Tk()
master.title('Student Registration System')
master.geometry("1250x700+210+100")
master.config(bg=background)

file = pathlib.Path("data.xlsx")

if file.exists():
    pass
else:
    file = workbook()
    sheet = file.active
    sheet['A1'] = "Registration No."
    sheet['B1'] = "Name"
    sheet['C1'] = "Class"
    sheet['D1'] = "Gender"
    sheet['E1'] = "DOB"
    sheet['F1'] = "Date of Registration"
    sheet['G1'] = "Religion"
    sheet['H1'] = "Skill"
    sheet['I1'] = "Father Name"
    sheet['J1'] = "Mother Name"
    sheet['K1'] = "Father's Occupation"
    sheet['L1'] = "Mother's Occupation"
    
    file.save('data.xlsx')
# exit the window    
def Exit():
    master.destroy()
    
#====================== Upload file =============================    
def upload():
    global photo_img3
    file_name = filedialog.askopenfilename(initialdir= os.getcwd(),
                                           title='select image file', filetypes=(('JPG files', '*.jpg'),
                                                                                 ('PNG files', '*.png'),
                                                                                 ('All files', '*.txt')))
    img = Image.open(file_name)
    resize_img = img.resize((290,285))
    photo_img3 = ImageTk.PhotoImage(resize_img)
    photo_label.config(image=photo_img3)
    photo_label.image = photo_img3
    
    
#======================== Save ==================================
def save():
    regis = registration.get()
    nam = name.get()
    cls = s_class.get()
    
    # if gender not select it will show error
    try:
        gen = gender
    except:
        messagebox.showerror('error','select gender')
        
    date_of_birth = dob.get()
    time_date = reg_date.get()
    reli = religion.get()
    ski = skill.get()
    f_name = father_name.get()
    m_name = mother_name.get()
    f_occupation = father_occupation.get()
    m_occupation = mother_occupation.get()
 
    # print(regis)
    # print(name)
    # print(cls)
    # print(date_of_birth)
    # print(time_date)
    # print(reli)
    # print(ski)
    # print(f_name)
    # print(m_name)
    # print(f_occupation)
    # print(m_occupation)
    
    # insert the data to the database
    if nam == '' or cls == '' or date_of_birth == '' or time_date == '' or reli == '' or ski == '' or f_name == '' or m_name == '' or f_occupation == '' or m_occupation == '':
        messagebox.showerror('error', 'Some feild still empty!')
        
    else:
        file = openpyxl.load_workbook('data.xlsx')
        sheet = file.active
        sheet.cell(column = 1, row = sheet.max_row+1, value = regis)
        sheet.cell(column = 2, row = sheet.max_row, value = nam)
        sheet.cell(column = 3, row = sheet.max_row, value = cls)
        sheet.cell(column = 4, row = sheet.max_row, value = gender)
        sheet.cell(column = 5, row = sheet.max_row, value = date_of_birth)
        sheet.cell(column = 6, row = sheet.max_row, value = time_date)
        sheet.cell(column = 7, row = sheet.max_row, value = reli)
        sheet.cell(column = 8, row = sheet.max_row, value = ski)
        sheet.cell(column = 9, row = sheet.max_row, value = f_name)
        sheet.cell(column = 10, row = sheet.max_row, value = m_name)
        sheet.cell(column = 11, row = sheet.max_row, value = f_occupation)
        sheet.cell(column = 12, row = sheet.max_row, value = m_occupation)
        
        file.save(r'data.xlsx')
        
        try:
            photo_img3.save('student_image/'+str(regis)+'*.jpg')
        except:
            messagebox.showinfo('info','profile picture is not available!!!')
            
        messagebox.showinfo('info', 'Successfully Save Data')
        
        
        
        registration_no() # it will recheck registration no and reissue registration no
        
#========================= Reset Button ==========================    
def reset():
   global photo_img2
   name.set('')
   dob.set('')
   religion.set('')
   skill.set('')
   father_name.set('')
   mother_name.set('')
   father_occupation.set('')
   mother_occupation.set('')
   s_class.set("select class")
   registration_no()
   save_button.config(state= 'normal')
   
   img = Image.open('photo.png')
   resize_img = img.resize((290,285))
   photo_img = ImageTk.PhotoImage(resize_img)
   photo_label.config(image= photo_img)
   photo_label.image = photo_img
   
   photo_img2 = ''
   
   
#========================== Search Button ========================
def search_button():
    text = search.get()
    save_button.config(state='disable') # after clicking on search save button will disable so that no one can click on it
    
    file = openpyxl.load_workbook('data.xlsx')
    sheet = file.active
    
    for row in sheet.rows:
        #print(row)
        if row[0].value == int(text):
            #print(row[0])
            name = row[0]
            #print('line 162',str(name))
            #regis_no_position = str(name)[14:-1]
            regis_number = str(name)[16:-1]
            # print('line 166',regis_no_position)
            # print('line 167',regis_number)
            
    try:
        print(str(name))
    except:
        messagebox.showerror('error','Invalid registration number')
        
    # registration_no_position showing like A1, A2, A3,.... An
    # registration_number just showing number after A2 like 2,3,... An
    
    x1 = sheet.cell(row = int(regis_number), column= 1).value
    x2 = sheet.cell(row = int(regis_number), column= 2).value
    x3 = sheet.cell(row = int(regis_number), column= 3).value
    x4 = sheet.cell(row = int(regis_number), column= 4).value
    x5 = sheet.cell(row = int(regis_number), column= 5).value
    x6 = sheet.cell(row = int(regis_number), column= 6).value
    x7 = sheet.cell(row = int(regis_number), column= 7).value
    x8 = sheet.cell(row = int(regis_number), column= 8).value
    x9 = sheet.cell(row = int(regis_number), column= 9).value
    x10 = sheet.cell(row = int(regis_number), column= 10).value
    x11 = sheet.cell(row = int(regis_number), column= 11).value
    x12 = sheet.cell(row = int(regis_number), column= 12).value
    
    # print(x1)
    # print(x2)
    # print(x3)
    # print(x4)
    # print(x5)
    # print(x6)
    # print(x7)
    # print(x8)
    # print(x9)
    # print(x10)
    # print(x11)
    # print(x12)
     
    registration.set(x1)
    #print('line 204', x2)
    #name.set(x2.value)
    s_class.set(x3)
    
    if x4 == 'Female':
        radio_button_female.select()
    else:
        radio_button_male.select()
        
    dob.set(x5)
    reg_date.set(x6)
    religion.set(x7)
    skill.set(x8)
    father_name.set(x9)
    mother_name.set(x10)
    father_occupation.set(x11)
    mother_occupation.set(x12)
    
    # img = Image.open('student_image/'+ str(x1)+'.jgp')
    # resize_img = img.resize((290,285))
    # photo_img2 = ImageTk.PhotoImage(resize_img)
    # photo_label.config(image= photo_img2)
    # photo_label.image = photo_img2
    
#========================== Update butoon ========================
def update():
    regis = registration.get()
    nam = name.get()
    cls = s_class.get()
    selaction()  
    gen = gender  
    date_of_birth = dob.get()
    time_date = reg_date.get()
    reli = religion.get()
    ski = skill.get()
    f_name = father_name.get()
    m_name = mother_name.get()
    f_occupation = father_occupation.get()
    m_occupation = mother_occupation.get()
    
    file = openpyxl.load_workbook('data.xlsx')
    sheet = file.active
    
    for row in sheet.rows:
        if row[0].value == regis:
            name_up = row[0]
            print(str(name_up))
            reg_no_position = str(name_up)[14 : -1]
            reg_number = str(name_up)[16 : -1]
            # print('line 253',reg_no_position)
            # print('line 254',reg_number)
            
    sheet.cell(column= 1, row= int(reg_number), value= regis)
    sheet.cell(column= 2, row= int(reg_number), value= nam)
    sheet.cell(column= 3, row= int(reg_number), value= cls)
    sheet.cell(column= 4, row= int(reg_number), value= gen)
    sheet.cell(column= 5, row= int(reg_number), value= date_of_birth)
    sheet.cell(column= 6, row= int(reg_number), value= time_date)
    sheet.cell(column= 7, row= int(reg_number), value= reli)
    sheet.cell(column= 8, row= int(reg_number), value= ski)
    sheet.cell(column= 9, row= int(reg_number), value= f_name)
    sheet.cell(column= 10, row= int(reg_number), value= m_name)
    sheet.cell(column= 11, row= int(reg_number), value= f_occupation)
    sheet.cell(column= 12, row= int(reg_number), value= m_occupation)
    
    file.save(r'data.xlsx')
    
    # try:
    #     img.save('student_image'+str(regis)+'.jpg')
    # except:
    #     pass
    
    # messagebox.showinfo('update', 'update successfully')
    
    
 
    
   
#=========================== Registration no =====================
def registration_no():
    file = openpyxl.load_workbook('data.xlsx')
    sheet = file.active
    row = sheet.max_row
    max_row_value = sheet.cell(row=row, column= 1).value
    
    try:
        registration.set(max_row_value+1)
    except:
        registration.set('1')

    

#========================== Gender function =======================   
def selaction():
    global gender
    value = radio.get()
    if value == 1:
        gender = 'Male'
    else:
        gender = 'Female'
    
# top frames
Label(master, text='Email: sagorluc@gmail.com', width=10, height=3, bg='#f0687c', anchor='e').pack(side=TOP, fill= X)
Label(master, text='STUDENT REGISTRATION', width=10, height=2, bg='#c36464', fg='#fff', font=('arial 20 bold')).pack(side=TOP, fill= X)

# search box to update
search = StringVar()
Entry(master, textvariable= search, width=15,bd=2, font=('arial 20')).place(x= 820, y= 70)


# img = Image.open('cloud.png')
# rsz_img = img.resize((56,30))
# photo_img = ImageTk.PhotoImage(rsz_img)
srch = Button(master, text='Search', compound=LEFT, bg='#68ddfa', font=('arial 13 bold'), command= search_button)
srch.place(x= 1060, y=70)

# add layer
img = Image.open('layer.png')
rsz_img = img.resize((20,20))
photo_img_layer = ImageTk.PhotoImage(rsz_img)
update_button = Button(master, image=photo_img_layer,bg='#c36464', command= update )
update_button.place(x= 110, y= 64)

#Registration and date
Label(master, text='Registration No:', font='arial 13', fg= framebg, bg=background).place(x=30, y=150)
Label(master, text='Date:', font='arial 13', fg= framebg, bg=background).place(x=500, y=150)

registration = IntVar()
reg_date = StringVar()

reg_entry = Entry(master, textvariable= registration, font='arial 10', width=15)
reg_entry.place(x=160, y=150)

registration_no()  # call the registration_no function

# date
today = date.today()
d1 = today.strftime('%d-%m-%y')

reg_date_entry = Entry(master, textvariable= reg_date, font='arial 10', width=15)
reg_date_entry.place(x=550, y=150)

reg_date.set(d1)

# student details
label_freme_student = LabelFrame(master, text="Student's details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=250, relief=GROOVE)
label_freme_student.place(x=30, y=200)

Label(label_freme_student, text='Full Name: ', font='arial 13', bg= framebg, fg= framefg).place(x=30, y=50)
Label(label_freme_student, text='Date of Birth: ', font='arial 13', bg= framebg, fg= framefg).place(x=30, y=100)
Label(label_freme_student, text='Gender: ', font='arial 13', bg= framebg, fg= framefg).place(x=30, y=150)

Label(label_freme_student, text='Class: ', font='arial 13', bg= framebg, fg= framefg).place(x=500, y=50)
Label(label_freme_student, text='Religion: ', font='arial 13', bg= framebg, fg= framefg).place(x=500, y=100)
Label(label_freme_student, text='Skill: ', font='arial 13', bg= framebg, fg= framefg).place(x=500, y=150)

# name entry input
global name
name = StringVar()
name_entry = Entry(label_freme_student, textvariable= name, width=20, font='arial 10')
name_entry.place(x=160, y=50)

dob = StringVar()
dob_entry = Entry(label_freme_student, textvariable= dob, width=20, font='arial 10')
dob_entry.place(x=160, y=100)


radio = IntVar()
radio_button_male = Radiobutton(label_freme_student, text='Male', variable= radio, value=1, bg=framebg, fg= framefg,command= selaction)
radio_button_male.place(x=150, y=150)

radio_button_female = Radiobutton(label_freme_student, text='Female', variable= radio, value=2, bg=framebg, fg= framefg, command= selaction)
radio_button_female.place(x=200, y=150)

religion = StringVar()
religion_entry = Entry(label_freme_student, textvariable= religion, width=20, font='arial 10')
religion_entry.place(x=630, y=100)

skill = StringVar()
skill_entry = Entry(label_freme_student, textvariable= skill, width=20, font='arial 10')
skill_entry.place(x=630, y=150)

s_class = Combobox(label_freme_student, values=[_ for _ in range(1,13)], font='Roboto 10',width=17, state='r')
s_class.place(x=630, y=50)
s_class.set('select class')

# parent details
label_freme_parent = LabelFrame(master, text="Parent's details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=220, relief=GROOVE)
label_freme_parent.place(x=30, y=470)

father_label = Label(label_freme_parent, text="Father's Name:", font='arial 10', bg= framebg, fg= framefg)
father_label.place(x=30, y=50)

mother_label = Label(label_freme_parent, text="Mother's Name:", font='arial 10', bg= framebg, fg= framefg)
mother_label.place(x=30, y=100)

father_label_occupation = Label(label_freme_parent, text="Father's Occupation:", font='arial 10', bg= framebg, fg= framefg)
father_label_occupation.place(x=500, y=50)

mother_label_occupation = Label(label_freme_parent, text="Mother's Occupation:", font='arial 10', bg= framebg, fg= framefg)
mother_label_occupation.place(x=500, y=100)

# father's and mother's all input
father_name = StringVar()
mother_name = StringVar()
father_name_entry = Entry(label_freme_parent, textvariable= father_name, width= 20, font='arial 10')
father_name_entry.place(x=160, y=50) 
mother_name_entry = Entry(label_freme_parent, textvariable= mother_name, width= 20, font='arial 10')
mother_name_entry.place(x=160, y=100)


father_occupation = StringVar()
mother_occupation = StringVar()
father_occupation_entry = Entry(label_freme_parent, textvariable= father_occupation, width= 20, font='arial 10')
father_occupation_entry.place(x=630, y=50) 
mother_occupation_entry = Entry(label_freme_parent, textvariable= mother_occupation, width= 20, font='arial 10')
mother_occupation_entry.place(x=630, y=100)

# uploead photo 
upload_frame = Frame(master, bd=3, bg='black', width=200, height=200, relief=GROOVE)
upload_frame.place(x=1000, y=160)

img = Image.open('photo.png')
resize_img = img.resize((290,285))
photo_img2 = ImageTk.PhotoImage(resize_img)

photo_label = Label(upload_frame, image= photo_img2, bg='black')
photo_label.place(x=-50, y=-15)

upload_button = Button(master, text='Upload', width=19, height=2, font='arial 12 bold', bg='lightblue', command= upload)
upload_button.place(x=1000, y=370)
save_button = Button(master, text='Save', width=19, height=2, font='arial 12 bold', bg='lightgreen', command= save)
save_button.place(x=1000, y=450)
reset_button = Button(master, text='Reset', width=19, height=2, font='arial 12 bold', bg='pink', command= reset)
reset_button.place(x=1000, y=530)
edit_button = Button(master, text='Exit', width=19, height=2, font='arial 12 bold', bg='grey', command= Exit)
edit_button.place(x=1000, y=610)

 
master.mainloop()
